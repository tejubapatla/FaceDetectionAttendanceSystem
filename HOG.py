import re
import cv2
import face_recognition
import numpy as np
from openpyxl import load_workbook

# from openpyxl import Workbook

filename = '1.jpg'
wb = load_workbook(filename='test.xlsx')
# ws2 = wb.create_sheet(title="atten")
sheet = wb['Pi']
names = ["ajaykumar", "gautham", "gayatri", "jagadeesh", "jahnavi", "jeevani", "pravallika", "premkumar", "tejaswi",
         "yeswanth"]
row = 2
col = 1
for name in names:
    sheet.cell(column=col, row=row, value=name)
    row += 1
row = 1
col = 2
for x in range(31):
    sheet.cell(column=col, row=row, value=str(x + 1) + "-01-2019")
    # worksheet.write(row, col,str(x+1)+"-01-2019")
    col += 1
# wb = Workbook()
# ws = wb.active

path = 'C:\\Users\\Teju\\Desktop\\WPy-3670\\settings\\projects\\finalupload\\server\\UwAmp\\www\\android_login_api\\up\\uploads\\'
filename = path + filename
with open('encodings.txt', 'r') as myfile:
    textData = myfile.read().replace('\n', '')
splitdata = textData.split(",")
with open('encodingnames.txt', 'r') as myfile:
    textNames = myfile.read().replace('\n', '')
names = textNames.split(",")

[float(i) for i in splitdata]
a = np.float64(splitdata)
feature = np.array_split(a, a.shape[0] / 128)

new_picture = face_recognition.load_image_file(filename)
image = cv2.imread(filename)
rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
boxes = face_recognition.face_locations(rgb, model='hog')
count = 0
for face_encoding in face_recognition.face_encodings(new_picture):
    counte = 4
    deviation = 0.62
    while counte > 1:
        results = face_recognition.compare_faces(feature, face_encoding, deviation)
        counte = results.count(True)
        deviation = deviation - 0.02

    print(counte)
    # cv2.rectangle(rgb, (boxes[count][1], boxes[count][0]), (boxes[count][3],boxes[count][2]), (0, 255, 0), 2)
    name = ''
    if True in results:
        first_match_index = results.index(True)
        name = names[first_match_index]

    if not name == '':
        cv2.rectangle(rgb, (boxes[count][1], boxes[count][0]), (boxes[count][3], boxes[count][2]), (255, 0, 0), 2)
        cv2.putText(rgb, name, (boxes[count][3] + 10, boxes[count][0]), 0, 0.5, (0, 255, 0), 2, cv2.LINE_4)
        sheet.cell(column=int(re.search(r'\d+', filename[len(filename) - 5:len(filename)]).group()) + 1,
                   row=first_match_index + 2, value="present")
        # worksheet.write(first_match_index+1, int(re.search(r'\d+',  filename[len(filename)-5:len(filename)]).group()),"present")
        # print(filename+'detected--'+str(int(re.search(r'\d+', filename[len(filename)-5:len(filename)]).group())))
        print('detected')
    else:
        print('detect failed')

    count = count + 1
wb.save(filename='test.xlsx')
bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
cv2.imwrite('detected.jpg', bgr)
cv2.imshow('image', bgr)
cv2.waitKey(0)
cv2.destroyAllWindows()
